import openpyxl

def duplicate_serial_number_column(excel_path, output_path):
    """
    在Excel文件的每个工作表中，在“释义”列前新增一列，内容复制序号列
    
    Args:
        excel_path (str): 原始Excel文件路径
        output_path (str): 处理后Excel文件保存路径
    """
    try:
        # 加载Excel文件
        wb = openpyxl.load_workbook(excel_path)
        print(f"成功加载文件，共包含 {len(wb.sheetnames)} 个工作表")
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"正在处理工作表: {sheet_name}")
            
            # 在第4列（释义列）前插入一列（新列成为第4列）
            ws.insert_cols(4)
            
            # 先设置新增列的表头（保持和序号列一致）
            ws.cell(row=1, column=4, value=ws.cell(row=1, column=1).value)
            
            # 遍历所有行，复制序号列数据到新增列
            # 从第2行开始（跳过表头），直到最后一行有数据的行
            for row in range(2, ws.max_row + 1):
                # 读取序号列（第1列）的数据
                serial_number = ws.cell(row=row, column=1).value
                # 写入新增的第4列
                ws.cell(row=row, column=4, value=serial_number)
        
        # 保存处理后的文件
        wb.save(output_path)
        wb.close()
        print(f"处理完成！文件已保存至: {output_path}")
        
    except FileNotFoundError:
        print(f"错误：未找到文件 {excel_path}")
    except Exception as e:
        print(f"处理过程中出现错误: {str(e)}")

# ------------------- 配置使用 -------------------
# 替换为你的原始Excel文件路径
input_file = "4_30.xlsx"
# 替换为处理后文件的保存路径（不要和原文件同名，避免覆盖）
output_file = "4_30_add_oneline.xlsx"

# 执行处理函数
duplicate_serial_number_column(input_file, output_file)